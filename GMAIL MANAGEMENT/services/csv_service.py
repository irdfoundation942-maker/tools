"""CSV/XLSX parsing and export helpers."""

from __future__ import annotations

import csv
import io
from typing import Iterable

REQUIRED_COLUMNS = ("Email",)
OPTIONAL_COLUMNS = ("Subject", "Message_Body")


class CSVValidationError(ValueError):
    pass


def _rows_from_dicts(dict_rows: Iterable[dict], field_map: dict) -> list[dict]:
    """Convert iterator of raw row dicts into normalized campaign rows."""
    missing = [col for col in REQUIRED_COLUMNS if col not in field_map]
    if missing:
        raise CSVValidationError(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Expected at minimum: {', '.join(REQUIRED_COLUMNS)} "
            f"(optional: {', '.join(OPTIONAL_COLUMNS)})."
        )

    rows: list[dict] = []
    for i, raw_row in enumerate(dict_rows, start=2):  # header is row 1
        email = (raw_row.get(field_map["Email"]) or "")
        email = email.strip() if isinstance(email, str) else str(email).strip()
        if not email:
            continue
        subject = raw_row.get(field_map["Subject"]) or "" if "Subject" in field_map else ""
        subject = subject.strip() if isinstance(subject, str) else str(subject).strip()
        body = raw_row.get(field_map["Message_Body"]) or "" if "Message_Body" in field_map else ""
        if not isinstance(body, str):
            body = str(body)
        rows.append({
            "Email": email,
            "Subject": subject,
            "Message_Body": body,
            "_row": i,
        })

    if not rows:
        raise CSVValidationError("No valid rows found.")
    return rows


def parse_campaign_csv(file_stream) -> list[dict]:
    """Parse an uploaded CSV into a list of row dicts.

    Accepts UTF-8 with or without BOM. Trims whitespace. Validates headers.
    """
    raw = file_stream.read()
    if isinstance(raw, bytes):
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
    else:
        text = raw

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise CSVValidationError("CSV appears to be empty.")

    field_map = {name.strip(): name for name in reader.fieldnames if name}
    return _rows_from_dicts(reader, field_map)


def parse_campaign_xlsx(file_stream) -> list[dict]:
    """Parse .xlsx/.xls into campaign rows using the first sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise CSVValidationError(
            "Excel support requires the 'openpyxl' package. "
            "Install it with: pip install openpyxl"
        ) from e

    try:
        wb = load_workbook(file_stream, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise CSVValidationError(f"Could not read Excel file: {e}") from e

    ws = wb.active
    if ws is None:
        raise CSVValidationError("Excel file has no sheets.")

    header_row = None
    row_iter = ws.iter_rows(values_only=True)
    for row in row_iter:
        if any(cell not in (None, "") for cell in row):
            header_row = row
            break
    if not header_row:
        raise CSVValidationError("Excel sheet appears to be empty.")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    field_map = {h: h for h in headers if h}

    def dict_iter():
        for row in row_iter:
            yield {headers[i]: row[i] for i in range(len(headers)) if headers[i]}

    return _rows_from_dicts(dict_iter(), field_map)


def parse_campaign_file(filename: str, file_stream) -> list[dict]:
    """Dispatch to CSV or Excel parser based on extension."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return parse_campaign_xlsx(file_stream)
    if name.endswith(".xls"):
        raise CSVValidationError(
            "Legacy .xls isn't supported — please save as .xlsx or .csv and re-upload."
        )
    return parse_campaign_csv(file_stream)


def contacts_to_csv(contacts: Iterable[dict]) -> str:
    """Serialize inbox contacts to a CSV string."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["name", "email", "subject", "date"])
    writer.writeheader()
    for c in contacts:
        writer.writerow({
            "name": c.get("name", ""),
            "email": c.get("email", ""),
            "subject": c.get("subject", ""),
            "date": c.get("date", ""),
        })
    return buf.getvalue()
